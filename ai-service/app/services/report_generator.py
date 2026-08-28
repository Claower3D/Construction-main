"""
QazGost AI — PDF & Excel Report Generator

Generates professional construction estimate reports:
- PDF with company header, SNiP references, cost breakdown
- Excel with formulas, summary sheets, charts

Usage:
    from app.services.report_generator import ReportGenerator
    gen = ReportGenerator()
    pdf_bytes = gen.generate_pdf(estimate_data)
    xlsx_bytes = gen.generate_excel(estimate_data)
"""

import io
import datetime
from typing import Any, Dict, List, Optional
from pathlib import Path
from loguru import logger


# ─── PDF Generator ─────────────────────────────────────────────

def generate_pdf(estimate: Dict[str, Any]) -> bytes:
    """Generate PDF estimate report.
    
    Returns PDF as bytes, ready to send as response.
    Falls back to simple text-based PDF if reportlab is not available.
    """
    try:
        return _generate_pdf_reportlab(estimate)
    except ImportError:
        logger.warning("reportlab not installed, generating simple PDF")
        return _generate_pdf_simple(estimate)


def _generate_pdf_reportlab(estimate: Dict[str, Any]) -> bytes:
    """Generate PDF using reportlab (professional layout)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        HRFlowable
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title_RU', parent=styles['Title'], fontSize=18)
    heading_style = ParagraphStyle('Heading_RU', parent=styles['Heading2'], fontSize=14)
    normal_style = styles['Normal']

    elements = []

    # Header
    now = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    estimate_id = estimate.get('id', 'QG-' + datetime.datetime.now().strftime("%Y%m%d%H%M"))

    elements.append(Paragraph(f"СМЕТА № {estimate_id}", title_style))
    elements.append(Spacer(1, 3*mm))
    elements.append(Paragraph(f"QazGost AI — Автоматический расчёт", normal_style))
    elements.append(Paragraph(f"Дата: {now}", normal_style))
    elements.append(Spacer(1, 2*mm))
    elements.append(HRFlowable(width="100%", color=colors.grey))
    elements.append(Spacer(1, 5*mm))

    # Object info
    city = estimate.get('city', '—')
    area = estimate.get('area_m2', 0)
    elements.append(Paragraph("Информация об объекте", heading_style))
    info_data = [
        ['Город', city.title()],
        ['Площадь', f'{area} м²'],
        ['Тип расчёта', estimate.get('type', 'Полная смета')],
    ]
    info_table = Table(info_data, colWidths=[60*mm, 100*mm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.Color(0.95, 0.95, 0.95)),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 8*mm))

    # Systems breakdown
    systems = estimate.get('systems', [])
    for sys in systems:
        sys_name = sys.get('system', 'Система')
        sys_total = sys.get('total_cost', 0)
        snip = sys.get('snip_code', '')

        elements.append(Paragraph(f"{sys_name} ({snip})", heading_style))
        elements.append(Spacer(1, 2*mm))

        # Items table
        items = sys.get('items', [])
        if items:
            table_data = [['№', 'Наименование', 'Объём', 'Ед.', 'Цена', 'Сумма']]
            for i, item in enumerate(items, 1):
                table_data.append([
                    str(i),
                    item.get('name', ''),
                    str(item.get('volume', '')),
                    item.get('unit', ''),
                    f"{item.get('unit_price', 0):,.0f}",
                    f"{item.get('total', 0):,.0f} ₸",
                ])

            # Subtotal row
            table_data.append(['', '', '', '', 'ИТОГО:', f"{sys_total:,.0f} ₸"])

            t = Table(table_data, colWidths=[10*mm, 60*mm, 20*mm, 15*mm, 25*mm, 30*mm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.2, 0.3)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('PADDING', (0, 0), (-1, -1), 4),
                ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.9, 0.95, 0.9)),
                ('FONTSIZE', (0, -1), (-1, -1), 10),
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ]))
            elements.append(t)

        elements.append(Spacer(1, 6*mm))

    # Grand total
    grand_total = estimate.get('grand_total', 0)
    nds = grand_total * 0.12
    total_with_nds = grand_total + nds

    elements.append(HRFlowable(width="100%", color=colors.black, thickness=2))
    elements.append(Spacer(1, 3*mm))

    total_data = [
        ['Итого без НДС:', f'{grand_total:,.0f} ₸'],
        ['НДС (12%):', f'{nds:,.0f} ₸'],
        ['ИТОГО с НДС:', f'{total_with_nds:,.0f} ₸'],
    ]
    total_table = Table(total_data, colWidths=[120*mm, 40*mm])
    total_table.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('FONTSIZE', (0, -1), (-1, -1), 14),
        ('PADDING', (0, 0), (-1, -1), 4),
        ('LINEBELOW', (0, -1), (-1, -1), 2, colors.black),
    ]))
    elements.append(total_table)

    elements.append(Spacer(1, 10*mm))
    elements.append(Paragraph(
        "Расчёт выполнен системой QazGost AI на основании действующих нормативов РК.",
        normal_style
    ))

    doc.build(elements)
    return buf.getvalue()


def _generate_pdf_simple(estimate: Dict[str, Any]) -> bytes:
    """Simple text-based PDF fallback (no reportlab)."""
    lines = []
    lines.append(f"СМЕТА — QazGost AI")
    lines.append(f"Дата: {datetime.datetime.now().strftime('%d.%m.%Y')}")
    lines.append(f"Город: {estimate.get('city', '—')}")
    lines.append(f"Площадь: {estimate.get('area_m2', 0)} м²")
    lines.append("")

    for sys in estimate.get('systems', []):
        lines.append(f"--- {sys.get('system', '')} ---")
        for item in sys.get('items', []):
            lines.append(f"  {item.get('name', '')}: {item.get('total', 0):,.0f} ₸")
        lines.append(f"  Итого: {sys.get('total_cost', 0):,.0f} ₸")
        lines.append("")

    grand = estimate.get('grand_total', 0)
    lines.append(f"ИТОГО без НДС: {grand:,.0f} ₸")
    lines.append(f"НДС (12%): {grand * 0.12:,.0f} ₸")
    lines.append(f"ИТОГО с НДС: {grand * 1.12:,.0f} ₸")

    text = "\n".join(lines)

    # Minimal PDF structure
    content = text.encode('utf-8')
    stream = f"""1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj
2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj
3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 595 842]/Contents 4 0 R/Resources<</Font<</F1 5 0 R>>>>>>endobj
5 0 obj<</Type/Font/Subtype/Type1/BaseFont/Courier>>endobj
4 0 obj<</Length {len(content) + 50}>>
stream
BT /F1 10 Tf 50 800 Td ({text}) Tj ET
endstream
endobj"""

    pdf = f"""%PDF-1.4
{stream}
xref
0 6
trailer<</Size 6/Root 1 0 R>>
startxref
0
%%EOF"""
    return pdf.encode('latin-1', errors='replace')


# ─── Excel Generator ───────────────────────────────────────────

def generate_excel(estimate: Dict[str, Any]) -> bytes:
    """Generate Excel estimate report with formulas.
    
    Returns XLSX as bytes.
    """
    try:
        return _generate_excel_openpyxl(estimate)
    except ImportError:
        logger.warning("openpyxl not installed, cannot generate Excel")
        raise ImportError("openpyxl required for Excel generation: pip install openpyxl")


def _generate_excel_openpyxl(estimate: Dict[str, Any]) -> bytes:
    """Generate Excel using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ─── Summary sheet ──────────────────────────────
    ws = wb.active
    ws.title = "Сводная смета"

    header_font = Font(bold=True, size=14)
    sub_font = Font(bold=True, size=11)
    money_fmt = '#,##0'
    header_fill = PatternFill(start_color="1a2744", end_color="1a2744", fill_type="solid")
    header_text = Font(bold=True, color="FFFFFF", size=10)
    total_fill = PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"СМЕТА — QazGost AI"
    ws['A1'].font = header_font

    ws['A2'] = 'Дата:'
    ws['B2'] = datetime.datetime.now().strftime('%d.%m.%Y')
    ws['A3'] = 'Город:'
    ws['B3'] = estimate.get('city', '—').title()
    ws['A4'] = 'Площадь:'
    ws['B4'] = f"{estimate.get('area_m2', 0)} м²"

    row = 6
    grand_total_row = None

    for sys in estimate.get('systems', []):
        sys_name = sys.get('system', '')
        snip = sys.get('snip_code', '')

        # System header
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"{sys_name} ({snip})"
        ws[f'A{row}'].font = sub_font
        row += 1

        # Table header
        headers = ['№', 'Наименование', 'Объём', 'Ед.', 'Цена за ед.', 'Сумма']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_text
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        row += 1

        # Items
        items = sys.get('items', [])
        first_item_row = row
        for i, item in enumerate(items, 1):
            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=item.get('name', '')).border = thin_border
            ws.cell(row=row, column=3, value=item.get('volume', 0)).border = thin_border
            ws.cell(row=row, column=4, value=item.get('unit', '')).border = thin_border

            price_cell = ws.cell(row=row, column=5, value=item.get('unit_price', 0))
            price_cell.number_format = money_fmt
            price_cell.border = thin_border

            # Formula: volume × unit_price
            formula_cell = ws.cell(row=row, column=6)
            formula_cell.value = f"=C{row}*E{row}"
            formula_cell.number_format = money_fmt
            formula_cell.border = thin_border

            row += 1

        # Subtotal
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = f"Итого {sys_name}:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = total_fill
        ws[f'A{row}'].border = thin_border

        subtotal_cell = ws.cell(row=row, column=6)
        if items:
            subtotal_cell.value = f"=SUM(F{first_item_row}:F{row-1})"
        else:
            subtotal_cell.value = sys.get('total_cost', 0)
        subtotal_cell.number_format = money_fmt
        subtotal_cell.font = Font(bold=True)
        subtotal_cell.fill = total_fill
        subtotal_cell.border = thin_border

        row += 2

    # Grand total section
    grand_total = estimate.get('grand_total', 0)

    ws[f'A{row}'] = 'Итого без НДС:'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.cell(row=row, column=6, value=grand_total).number_format = money_fmt
    ws[f'F{row}'].font = Font(bold=True, size=12)
    grand_total_row = row
    row += 1

    ws[f'A{row}'] = 'НДС (12%):'
    ws.cell(row=row, column=6).value = f"=F{grand_total_row}*0.12"
    ws[f'F{row}'].number_format = money_fmt
    nds_row = row
    row += 1

    ws[f'A{row}'] = 'ИТОГО С НДС:'
    ws[f'A{row}'].font = Font(bold=True, size=14, color="006400")
    ws.cell(row=row, column=6).value = f"=F{grand_total_row}+F{nds_row}"
    ws[f'F{row}'].number_format = money_fmt
    ws[f'F{row}'].font = Font(bold=True, size=14, color="006400")

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
